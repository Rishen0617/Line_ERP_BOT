"""Local xlsx storage — primary data layer.

All reads and writes go here first.
Google Sheets receives async fire-and-forget syncs (see sheets_service.py).

Thread-safety: a process-level asyncio.Lock serialises all writes so concurrent
BackgroundTasks cannot corrupt the xlsx file.
"""
from __future__ import annotations

import asyncio
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.config import settings

log = logging.getLogger(__name__)

# One lock per process — serialises all xlsx writes.
_WRITE_LOCK = asyncio.Lock()

# Tab name → header row (mirrors sheets_service sheet layout).
# 月報 is formula-driven; we never write to it programmatically.
SHEET_HEADERS: dict[str, list[str]] = {
    "收據記錄": [
        "日期", "上傳時間", "上傳者", "文件類型", "廠商/客戶",
        "品項摘要", "金額", "稅額", "單號", "圖片連結", "AI信心度", "群組", "備註",
    ],
    "記帳流水帳": ["日期", "類別", "說明", "金額", "對象", "關聯單號", "記錄者"],
    "訂單追蹤": ["訂單日期", "供應商", "品項摘要", "金額", "狀態", "物流單號", "建立者"],
    "班表": [
        "日期", "員工ID", "員工名", "店別",
        "開始", "結束", "工時", "班別", "狀態", "備註",
    ],
    "請假記錄": [
        "申請時間", "員工ID", "員工名", "假日",
        "假別", "原因", "狀態", "審核者", "備註",
    ],
    "電商訂單": [
        "建立時間", "訂單編號", "平台", "客戶", "電話", "地址", "品項",
        "商品金額", "運費", "總金額", "付款方式", "付款狀態",
        "物流公司", "物流單號", "出貨狀態", "備註", "建立者",
    ],
    "庫存台帳": ["品項", "規格", "單位", "安全庫存", "目前庫存", "最後更新", "分類", "供應商"],
    "庫存異動": [
        "時間", "品項", "異動類型", "數量", "單位",
        "店別", "關聯單號", "操作者", "備註",
    ],
    "廠商主檔": [
        "廠商名", "發票類型", "請款週期", "銀行代碼",
        "帳號", "戶名", "LINE群組", "備註",
    ],
    "應付帳款": [
        "到貨日期", "廠商名", "品項", "數量", "單位", "單價", "金額",
        "發票類型", "單號", "請款週期", "狀態", "匯款日期", "備註",
    ],
}


# ─── Internal file helpers ────────────────────────────────────────────


def _path() -> Path:
    return Path(settings.xlsx_path)


def _load_or_create() -> Workbook:
    """Load workbook from disk, or create a new one with correct sheet structure."""
    p = _path()
    if p.exists():
        wb = load_workbook(str(p))
    else:
        wb = Workbook()
        # Remove the default blank sheet openpyxl creates
        for name in wb.sheetnames:
            del wb[name]

    # Ensure all required tabs exist with their header rows.
    for sheet_name, headers in SHEET_HEADERS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            if headers:
                ws.append(headers)
            log.info("xlsx: created missing sheet '%s'", sheet_name)

    return wb


def _save(wb: Workbook) -> None:
    wb.save(str(_path()))


# ─── Public API ───────────────────────────────────────────────────────


async def append_row(sheet_name: str, row: list[Any]) -> int:
    """Append *row* to *sheet_name*. Returns the 1-based row number of the new row."""
    str_row = [str(v) if v is not None else "" for v in row]

    async with _WRITE_LOCK:
        def _do() -> int:
            wb = _load_or_create()
            ws = wb[sheet_name]
            ws.append(str_row)
            row_num = ws.max_row
            _save(wb)
            return row_num

        result = await asyncio.to_thread(_do)

    log.info("xlsx append [%s] → row %d", sheet_name, result)
    return result


async def read_all_rows(sheet_name: str) -> list[list[Any]]:
    """Return all data rows (header excluded) as list[list].

    Cells with None values are normalised to empty string.
    """
    def _do() -> list[list[Any]]:
        p = _path()
        if not p.exists():
            return []
        wb = load_workbook(str(p), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]
        rows: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # skip header
                continue
            rows.append(["" if v is None else v for v in row])
        wb.close()
        return rows

    return await asyncio.to_thread(_do)


async def update_cells(
    sheet_name: str,
    row_num: int,
    updates: dict[int, Any],
) -> None:
    """Update specific cells in *row_num* (1-based).

    *updates* maps 0-based column index → new value.
    """
    async with _WRITE_LOCK:
        def _do() -> None:
            wb = _load_or_create()
            ws = wb[sheet_name]
            for col_0, value in updates.items():
                ws.cell(
                    row=row_num,
                    column=col_0 + 1,
                    value="" if value is None else str(value),
                )
            _save(wb)

        await asyncio.to_thread(_do)

    log.info("xlsx update [%s] row %d cols %s", sheet_name, row_num, list(updates.keys()))


async def find_row_by_col(
    sheet_name: str,
    col_idx: int,
    value: str,
) -> int | None:
    """Return the 1-based row number of the first row where column *col_idx* (0-based)
    equals *value*, or None if not found. Header row is row 1; data starts at row 2.
    """
    rows = await read_all_rows(sheet_name)
    for i, row in enumerate(rows):
        cell = str(row[col_idx]) if len(row) > col_idx else ""
        if cell == value:
            return i + 2  # +1 for 0-based→1-based, +1 for skipped header
    return None


async def read_row(sheet_name: str, row_num: int) -> list[Any]:
    """Return a single row by 1-based row number (header = row 1)."""
    def _do() -> list[Any]:
        p = _path()
        if not p.exists():
            return []
        wb = load_workbook(str(p), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]
        row_data = [
            "" if cell.value is None else cell.value
            for cell in ws[row_num]
        ]
        wb.close()
        return row_data

    return await asyncio.to_thread(_do)
